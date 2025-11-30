# ui.py
import tkinter as tk
from tkinter import ttk, messagebox, Toplevel, filedialog
import threading
import queue
import time
import csv
import os
import random
import shutil
import base64
import uuid
from browser import BrowserManager
from logger import log_report, migrate_log_if_needed
import config

class ReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title(config.WINDOW_TITLE)
        self.root.geometry("1200x750")
        
        # Migrate log file if needed
        migrate_log_if_needed()
        
        # State
        self.is_running = False
        self.stop_event = threading.Event()
        self.active_browsers = {} # {account_id: browser_instance}
        
        # Data Model
        self.all_accounts = [] # List of dicts: {id, c_user, xs, cookie, status, result}
        self.account_map = {} # {id: dict} for O(1) lookup
        self.filtered_accounts = []
        self.current_page = 0
        self.page_size = 50
        
        # self.final_screenshots = {} # REMOVED: Save to disk instead
        self.report_sets = [] # List of (category, detail) tuples
        self.lock = threading.Lock()
        
        # UI Queue for performance
        self.ui_queue = queue.Queue()
        
        # Temp folder for screenshots
        self.temp_dir = "temp_screenshots"
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
        
        self.setup_ui()
        
        # Start UI consumer
        self.process_ui_queue()
        
        # Batch timing & counters
        self.batch_start_time = None
        self.processed_count = 0
        self.success_count = 0
        self.failure_count = 0
        self.total_accounts_in_batch = 0

    def process_ui_queue(self):
        try:
            # Process more events to drain queue faster
            for _ in range(200):
                task = self.ui_queue.get_nowait()
                func, args = task
                try:
                    func(*args)
                except Exception:
                    pass
                self.ui_queue.task_done()
        except queue.Empty:
            pass
        # Check again sooner
        self.root.after(20, self.process_ui_queue)

    def setup_ui(self):
        # Setup UI
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill='both', expand=True)

        # Use a PanedWindow so panels are resizable and the right panel won't be hidden
        paned = ttk.PanedWindow(main_frame, orient='horizontal')
        paned.pack(fill='both', expand=True)
        self.paned = paned
        
        # === LEFT PANEL: ACCOUNT LIST & SETTINGS ===
        left_frame = ttk.LabelFrame(paned, text="1. Quản lý Tài khoản (Cookies)", padding=10)
        # Add left and right frames to paned window. Left gets more weight.
        paned.add(left_frame, weight=3)
        # Ensure left pane keeps a minimum size so its controls are not hidden
        try:
            paned.paneconfigure(left_frame, minsize=450)
        except Exception:
            try:
                left_frame.config(width=450)
            except Exception:
                pass
        
        # Control Bar for List
        control_frame = ttk.Frame(left_frame)
        control_frame.pack(fill='x', pady=5)
        
        # Cookie Input (c_user & xs)
        cookie_frame = ttk.Frame(control_frame)
        cookie_frame.pack(fill='x', pady=2)
        
        ttk.Label(cookie_frame, text="c_user:").pack(side='left')
        self.entry_c_user = ttk.Entry(cookie_frame, width=15)
        self.entry_c_user.pack(side='left', padx=5)
        
        ttk.Label(cookie_frame, text="xs:").pack(side='left')
        self.entry_xs = ttk.Entry(cookie_frame, width=15)
        self.entry_xs.pack(side='left', padx=5)
        
        ttk.Label(cookie_frame, text="Proxy:").pack(side='left', padx=(6,0))
        self.entry_cookie_proxy = ttk.Entry(cookie_frame, width=20)
        self.entry_cookie_proxy.pack(side='left', padx=5)

        ttk.Button(cookie_frame, text="Thêm Cookie", command=self.add_cookie).pack(side='left', padx=5)

        # File Actions
        file_frame = ttk.Frame(control_frame)
        file_frame.pack(fill='x', pady=5)
        
        ttk.Button(file_frame, text="Nhập Excel (CSV)", command=self.import_cookies).pack(side='left')
        ttk.Button(file_frame, text="Nhập (.xlsx)", command=self.import_cookies_xlsx).pack(side='left', padx=5)
        
        ttk.Button(file_frame, text="Xuất Excel (CSV)", command=self.export_cookies).pack(side='left')
        ttk.Button(file_frame, text="Xuất (.xlsx)", command=self.export_cookies_xlsx).pack(side='left', padx=5)
        
        ttk.Button(file_frame, text="Xóa chọn", command=self.delete_selected).pack(side='left')

        # Search & Filter
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill='x', pady=5)
        ttk.Label(search_frame, text="Tìm kiếm:").pack(side='left')
        self.entry_search = ttk.Entry(search_frame)
        self.entry_search.pack(side='left', fill='x', expand=True, padx=5)
        self.entry_search.bind("<KeyRelease>", self.on_search)
        
        self.combo_status_filter = ttk.Combobox(search_frame, values=["Tất cả", "Chờ", "Đang chạy...", "Hoàn thành", "Lỗi", "Lỗi Start"], state="readonly", width=15)
        self.combo_status_filter.current(0)
        self.combo_status_filter.pack(side='left', padx=5)
        self.combo_status_filter.bind("<<ComboboxSelected>>", self.on_search)

        # Treeview
        columns = ("stt", "cookie", "proxy", "status", "result", "view")
        self.tree = ttk.Treeview(left_frame, columns=columns, show='headings', selectmode='extended')
        self.tree.heading("stt", text="#")
        self.tree.heading("cookie", text="Cookie (Ẩn)")
        self.tree.heading("proxy", text="Proxy")
        self.tree.heading("status", text="Trạng thái")
        self.tree.heading("result", text="Kết quả")
        self.tree.heading("view", text="Hành động")
        
        self.tree.column("stt", width=40, anchor='center')
        self.tree.column("cookie", width=150)
        self.tree.column("proxy", width=120)
        self.tree.column("status", width=100)
        self.tree.column("result", width=200)
        self.tree.column("view", width=80, anchor='center')
        
        scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Pagination Controls
        page_frame = ttk.Frame(left_frame)
        page_frame.pack(fill='x', pady=5)
        ttk.Button(page_frame, text="<< Trước", command=lambda: self.change_page(-1)).pack(side='left')
        self.lbl_page = ttk.Label(page_frame, text="Trang 1/1")
        self.lbl_page.pack(side='left', padx=10)
        ttk.Button(page_frame, text="Sau >>", command=lambda: self.change_page(1)).pack(side='left')
        ttk.Label(page_frame, text="Tổng: 0").pack(side='right', padx=5)
        self.lbl_total = page_frame.winfo_children()[-1]

        # Bind click for View column
        self.tree.bind("<Button-1>", self.on_tree_click)
        
        # Context Menu for View Browser
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Xem trình duyệt (Screenshot)", command=self.view_browser_snapshot)

        # Settings Frame (Bottom Left) - pinned to bottom so it stays visible
        settings_frame = ttk.Frame(left_frame)
        settings_frame.pack(side='bottom', fill='x', pady=10)
        
        ttk.Label(settings_frame, text="Số luồng (Threads):").pack(side='left')
        self.spin_threads = ttk.Spinbox(settings_frame, from_=1, to=50, width=5)
        self.spin_threads.set(2)
        self.spin_threads.pack(side='left', padx=5)
        
        self.var_headless = tk.BooleanVar(value=True)
        ttk.Checkbutton(settings_frame, text="Chạy ngầm (Headless)", variable=self.var_headless).pack(side='left', padx=10)

        ttk.Label(settings_frame, text="Proxy chung:").pack(side='left', padx=(10,0))
        self.entry_proxy = ttk.Entry(settings_frame)
        # Make proxy entry expand so it remains visible when left pane narrows
        self.entry_proxy.pack(side='left', padx=5, fill='x', expand=True)

        # === RIGHT PANEL: REPORT CONFIG ===
        right_frame = ttk.LabelFrame(paned, text="2. Cấu hình Báo cáo", padding=10)
        # Keep a reasonable minimum width so the panel isn't fully hidden when window is small
        paned.add(right_frame, weight=1)
        # Set minimum size via paneconfigure to avoid unsupported add() option on some Tk versions
        try:
            paned.paneconfigure(right_frame, minsize=280)
        except Exception:
            # Fallback: set a fixed width on the frame so it doesn't disappear
            try:
                right_frame.config(width=280)
            except Exception:
                pass
        self.right_frame = right_frame
        
        ttk.Label(right_frame, text="Link cần báo cáo:").pack(anchor='w')
        self.entry_url = ttk.Entry(right_frame, width=40)
        self.entry_url.pack(fill='x', pady=5)

        ttk.Label(right_frame, text="Hạng mục chính:").pack(anchor='w', pady=(10,0))
        self.combo_category = ttk.Combobox(right_frame, values=config.CATEGORIES, state="readonly")
        self.combo_category.pack(fill='x', pady=5)
        self.combo_category.bind("<<ComboboxSelected>>", self.on_category_change)

        ttk.Label(right_frame, text="Chi tiết hành vi:").pack(anchor='w', pady=(10,0))
        self.combo_detail = ttk.Combobox(right_frame, state="readonly")
        self.combo_detail.pack(fill='x', pady=5)
        
        # Report Sets UI
        btn_set_frame = ttk.Frame(right_frame)
        btn_set_frame.pack(fill='x', pady=5)
        ttk.Button(btn_set_frame, text="Thêm vào bộ (Random)", command=self.add_report_set).pack(side='left', expand=True, fill='x')
        ttk.Button(btn_set_frame, text="Xóa bộ", command=self.clear_report_set).pack(side='left', padx=5)

        ttk.Label(right_frame, text="Danh sách bộ báo cáo:").pack(anchor='w')
        self.list_report_sets = tk.Listbox(right_frame, height=6)
        self.list_report_sets.pack(fill='x', pady=5)
        
        ttk.Separator(right_frame, orient='horizontal').pack(fill='x', pady=20)
        
        self.btn_run = ttk.Button(right_frame, text="CHẠY HÀNG LOẠT", command=self.start_batch)
        self.btn_run.pack(fill='x', ipady=10)
        
        self.btn_stop = ttk.Button(right_frame, text="DỪNG TẤT CẢ", command=self.stop_batch, state='disabled')
        self.btn_stop.pack(fill='x', pady=5)
        
        self.btn_history = ttk.Button(right_frame, text="Xem Lịch sử", command=self.view_history)
        self.btn_history.pack(fill='x', pady=5)
        
        self.lbl_status = ttk.Label(right_frame, text="Sẵn sàng", relief="sunken", anchor="w")
        self.lbl_status.pack(side='bottom', fill='x', pady=10)
        # Batch stats (Elapsed, Remaining, Success %)
        stats_frame = ttk.Frame(right_frame)
        stats_frame.pack(side='bottom', fill='x', pady=(0,6))

        self.lbl_elapsed = ttk.Label(stats_frame, text="Elapsed: 00:00:00")
        self.lbl_elapsed.pack(side='left')

        self.lbl_remaining = ttk.Label(stats_frame, text="Remaining: --:--:--")
        self.lbl_remaining.pack(side='left', padx=10)

        self.lbl_success = ttk.Label(stats_frame, text="Success: 0% (0/0)")
        self.lbl_success.pack(side='left', padx=10)

        # Init
        if config.CATEGORIES:
            self.combo_category.current(0)
            self.on_category_change()

    # --- EVENT HANDLERS ---
    def add_cookie(self):
        c_user = self.entry_c_user.get().strip()
        xs = self.entry_xs.get().strip()
        proxy_val = self.entry_cookie_proxy.get().strip()
        
        if c_user and xs:
            full_cookie = f"c_user={c_user};xs={xs}"
            display_c = f"{c_user} | {xs[:5]}..."
            
            new_acc = {
                "id": str(uuid.uuid4()),
                "c_user": c_user,
                "xs": xs,
                "proxy": proxy_val,
                "cookie": full_cookie,
                "display_c": display_c,
                "status": "Chờ",
                "result": ""
            }
            self.all_accounts.append(new_acc)
            self.account_map[new_acc['id']] = new_acc
            self.refresh_data()
            
            self.entry_c_user.delete(0, 'end')
            self.entry_xs.delete(0, 'end')
            self.entry_cookie_proxy.delete(0, 'end')
        else:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập cả c_user và xs")

    def refresh_data(self):
        # Filter
        search_txt = self.entry_search.get().lower()
        status_filter = self.combo_status_filter.get()
        
        filtered = []
        for acc in self.all_accounts:
            # Search text in c_user or result
            match_text = (search_txt in acc['c_user'].lower()) or (search_txt in acc['result'].lower())
            # Filter status
            match_status = True
            if status_filter != "Tất cả":
                if status_filter == "Lỗi" and "Lỗi" in acc['status']:
                    match_status = True
                else:
                    match_status = (acc['status'] == status_filter)
            
            if match_text and match_status:
                filtered.append(acc)
        
        self.filtered_accounts = filtered
        
        # Pagination
        total = len(self.filtered_accounts)
        max_page = (total - 1) // self.page_size
        if self.current_page > max_page: self.current_page = max_page
        if self.current_page < 0: self.current_page = 0
        
        start = self.current_page * self.page_size
        end = start + self.page_size
        page_items = self.filtered_accounts[start:end]
        
        # Update Treeview
        # Clear all
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        for i, acc in enumerate(page_items):
            # Use acc['id'] as iid if possible, but treeview iid must be unique in tree.
            # We use acc['id'] to map back.
            self.tree.insert("", "end", iid=acc['id'], values=(start + i + 1, acc['display_c'], acc.get('proxy', ''), acc['status'], acc['result'], "Xem"))
            
        self.lbl_page.config(text=f"Trang {self.current_page + 1}/{max_page + 1 if total > 0 else 1}")
        self.lbl_total.config(text=f"Tổng: {total}")

    def on_search(self, event=None):
        self.current_page = 0
        self.refresh_data()

    def change_page(self, delta):
        total = len(self.filtered_accounts)
        max_page = (total - 1) // self.page_size
        new_page = self.current_page + delta
        if 0 <= new_page <= max_page:
            self.current_page = new_page
            self.refresh_data()

    def import_cookies(self):
        fp = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")])
        if fp:
            try:
                with open(fp, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    if not header: return
                    
                    try:
                        headers = [h.strip().lower() for h in header]
                        idx_c = headers.index('c_user')
                        idx_xs = headers.index('xs')
                        idx_proxy = headers.index('proxy') if 'proxy' in headers else None
                    except ValueError:
                        messagebox.showerror("Lỗi Format", "File CSV cần có cột 'c_user' và 'xs'")
                        return
                        
                    count = 0
                    for row in reader:
                        if len(row) > max(idx_c, idx_xs):
                            c_user = row[idx_c].strip()
                            xs = row[idx_xs].strip()
                            if c_user and xs:
                                full_cookie = f"c_user={c_user};xs={xs}"
                                display_c = f"{c_user} | {xs[:5]}..."
                                proxy_val = row[idx_proxy].strip() if idx_proxy is not None and len(row) > idx_proxy else ""
                                new_acc = {
                                    "id": str(uuid.uuid4()),
                                    "c_user": c_user,
                                    "xs": xs,
                                    "proxy": proxy_val,
                                    "cookie": full_cookie,
                                    "display_c": display_c,
                                    "status": "Chờ",
                                    "result": ""
                                }
                                self.all_accounts.append(new_acc)
                                self.account_map[new_acc['id']] = new_acc
                                count += 1
                    self.refresh_data()
                    messagebox.showinfo("Thành công", f"Đã nhập {count} tài khoản.")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không đọc được file: {e}")

    def export_cookies(self):
        fp = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV Files", "*.csv")])
        if fp:
            try:
                with open(fp, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    # include proxy column
                    writer.writerow(["c_user", "xs", "proxy"])
                    for acc in self.all_accounts:
                        writer.writerow([acc['c_user'], acc['xs'], acc.get('proxy','')])
                messagebox.showinfo("Thành công", "Đã xuất file CSV.")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không ghi được file: {e}")

    def delete_selected(self):
        selected_ids = self.tree.selection()
        if not selected_ids: return
        
        # Remove from all_accounts
        # selected_ids contains the 'id' (uuid) because we set iid=acc['id']
        self.all_accounts = [acc for acc in self.all_accounts if acc['id'] not in selected_ids]
        # Rebuild map
        self.account_map = {acc['id']: acc for acc in self.all_accounts}
        
        self.refresh_data()
        
        # If no accounts left, cleanup temp screenshots folder
        try:
            if not self.all_accounts:
                if os.path.exists(self.temp_dir):
                    for fname in os.listdir(self.temp_dir):
                        fpath = os.path.join(self.temp_dir, fname)
                        try:
                            if os.path.isfile(fpath):
                                os.remove(fpath)
                        except Exception:
                            pass
        except Exception:
            pass

    def on_category_change(self, event=None):
        cat = self.combo_category.get()
        details = config.REPORT_DATA.get(cat, [])
        self.combo_detail.config(values=details)
        if details:
            self.combo_detail.current(0)
        else:
            self.combo_detail.set("")

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def on_tree_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            col = self.tree.identify_column(event.x)
            if col == "#6": # Column 'view'
                item_id = self.tree.identify_row(event.y)
                if item_id:
                    self.open_preview(item_id)

    def open_preview(self, item_id):
        # Find account in all_accounts (O(1))
        acc = self.account_map.get(item_id)
        if not acc: return
        
        status = acc['status']
        
        top = Toplevel(self.root)
        top.title(f"Xem: {acc['c_user']} - {status}")
        top.geometry("800x600")
        
        lbl_img = tk.Label(top, text="Đang tải hình ảnh...", bg="black", fg="white")
        lbl_img.pack(fill='both', expand=True)
        
        # If running, update live. If finished, show final.
        if status in ["Chờ", "Lỗi Start"]:
             lbl_img.config(text="Trình duyệt chưa chạy hoặc lỗi khởi động.")
        elif status in ["Hoàn thành", "Lỗi"]:
            # Show final from disk
            b64 = self.get_screenshot_from_disk(item_id)
            if b64:
                self.show_image(lbl_img, b64)
            else:
                lbl_img.config(text="Không có ảnh (Đã đóng).")
        else:
            # Running
            self.update_live_preview(top, lbl_img, item_id)

    def update_live_preview(self, top, lbl_img, item_id):
        if not top.winfo_exists(): return
        
        with self.lock:
            bm = self.active_browsers.get(item_id)
            
        if bm:
            b64 = bm.get_screenshot_base64()
            if b64:
                self.show_image(lbl_img, b64)
            else:
                pass 
        else:
            # Browser might have just finished, check disk
            b64 = self.get_screenshot_from_disk(item_id)
            if b64:
                self.show_image(lbl_img, b64)
            return

        # Refresh every 2s to reduce load
        top.after(2000, lambda: self.update_live_preview(top, lbl_img, item_id))

    def show_image(self, lbl, b64):
        try:
            img = tk.PhotoImage(data=b64)
            # Simple resize logic (subsample) if too large
            w = img.width()
            h = img.height()
            if w > 1000:
                scale = w // 800 + 1
                img = img.subsample(scale, scale)
            
            lbl.config(image=img, text="")
            lbl.image = img
        except Exception as e:
            lbl.config(text=f"Lỗi hiển thị ảnh: {e}")

    def view_browser_snapshot(self):
        sel = self.tree.selection()
        if not sel: return
        item_id = sel[0]
        self.open_preview(item_id)

    def view_history(self):
        if not os.path.exists(config.LOG_FILE):
             messagebox.showinfo("Info", "Chưa có lịch sử báo cáo.")
             return

        try:
            with open(config.LOG_FILE, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                data = list(reader)
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không đọc được file log: {e}")
            return

        if not data:
            messagebox.showinfo("Info", "File lịch sử trống.")
            return

        header = data[0]
        self.history_header = header
        rows = data[1:]
        # Reverse to show newest first
        rows.reverse()
        
        self.history_rows_all = rows # Store all rows for filtering
        self.history_rows = rows # Current filtered rows
        self.history_page = 0
        self.history_page_size = 20
        
        top = Toplevel(self.root)
        top.title("Lịch sử Báo cáo")
        top.geometry("1000x600")
        
        # Filter Frame
        filter_frame = ttk.Frame(top)
        filter_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(filter_frame, text="Tìm kiếm (URL, Account, Kết quả...):").pack(side='left')
        self.entry_hist_search = ttk.Entry(filter_frame)
        self.entry_hist_search.pack(side='left', fill='x', expand=True, padx=5)
        self.entry_hist_search.bind("<KeyRelease>", self.filter_history)

        # Treeview Frame
        tree_frame = ttk.Frame(top)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Scrollbars
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        self.hist_tree = ttk.Treeview(tree_frame, columns=header, show='headings', 
                                      yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        scrollbar_y.config(command=self.hist_tree.yview)
        scrollbar_x.config(command=self.hist_tree.xview)
        
        scrollbar_y.pack(side='right', fill='y')
        scrollbar_x.pack(side='bottom', fill='x')
        self.hist_tree.pack(side='left', fill='both', expand=True)
        
        for col in header:
            self.hist_tree.heading(col, text=col)
            self.hist_tree.column(col, width=150, minwidth=100)
            
        # Pagination Controls
        ctrl_frame = ttk.Frame(top, padding=10)
        ctrl_frame.pack(fill='x', side='bottom')
        
        ttk.Button(ctrl_frame, text="<< Trước", command=lambda: self.change_hist_page(-1)).pack(side='left')
        self.lbl_page_info = ttk.Label(ctrl_frame, text="Trang 1")
        self.lbl_page_info.pack(side='left', padx=20)
        ttk.Button(ctrl_frame, text="Sau >>", command=lambda: self.change_hist_page(1)).pack(side='left')
        
        ttk.Button(ctrl_frame, text="Xuất Excel (.xlsx)", command=self.export_history_xlsx).pack(side='right', padx=10)
        
        self.load_hist_page()

    def filter_history(self, event=None):
        keyword = self.entry_hist_search.get().lower()
        if not keyword:
            self.history_rows = list(self.history_rows_all)
        else:
            # Filter if keyword appears in any column of the row
            self.history_rows = [
                row for row in self.history_rows_all 
                if any(keyword in str(cell).lower() for cell in row)
            ]
        self.history_page = 0
        self.load_hist_page()

    def change_hist_page(self, delta):
        new_page = self.history_page + delta
        total_rows = len(self.history_rows)
        if total_rows == 0: return
        max_page = (total_rows - 1) // self.history_page_size
        
        if 0 <= new_page <= max_page:
            self.history_page = new_page
            self.load_hist_page()

    def load_hist_page(self):
        # Clear current items
        for item in self.hist_tree.get_children():
            self.hist_tree.delete(item)
            
        start = self.history_page * self.history_page_size
        end = start + self.history_page_size
        page_data = self.history_rows[start:end]
        
        for row in page_data:
            self.hist_tree.insert("", "end", values=row)
            
        total_rows = len(self.history_rows)
        if total_rows == 0:
            max_page = 1
        else:
            max_page = (total_rows - 1) // self.history_page_size + 1
        self.lbl_page_info.config(text=f"Trang {self.history_page + 1} / {max_page} (Tổng: {total_rows} dòng)")

    def import_cookies_xlsx(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Thiếu thư viện", "Vui lòng cài đặt openpyxl: pip install openpyxl")
            return

        fp = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if fp:
            try:
                wb = openpyxl.load_workbook(fp)
                sheet = wb.active
                
                # Find headers
                headers = {}
                for cell in sheet[1]:
                    if cell.value:
                        headers[str(cell.value).strip().lower()] = cell.column - 1
                
                if 'c_user' not in headers or 'xs' not in headers:
                    messagebox.showerror("Lỗi Format", "File Excel cần có cột 'c_user' và 'xs' ở dòng đầu tiên.")
                    return
                
                idx_c = headers['c_user']
                idx_xs = headers['xs']
                
                count = 0
                # Determine if there's a proxy column in headers (header keys map to indices)
                proxy_idx = None
                # headers dict maps lowercased header -> index, we can check
                for k, v in headers.items():
                    if k == 'proxy':
                        proxy_idx = v
                        break

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        if row[idx_c] and row[idx_xs]:
                            c_user = str(row[idx_c]).strip()
                            xs = str(row[idx_xs]).strip()

                            full_cookie = f"c_user={c_user};xs={xs}"
                            display_c = f"{c_user} | {xs[:5]}..."
                            proxy_val = ''
                            if proxy_idx is not None and len(row) > proxy_idx and row[proxy_idx] is not None:
                                proxy_val = str(row[proxy_idx]).strip()

                            new_acc = {
                                "id": str(uuid.uuid4()),
                                "c_user": c_user,
                                "xs": xs,
                                "proxy": proxy_val,
                                "cookie": full_cookie,
                                "display_c": display_c,
                                "status": "Chờ",
                                "result": ""
                            }
                            self.all_accounts.append(new_acc)
                            self.account_map[new_acc['id']] = new_acc
                            count += 1
                    except Exception:
                        continue
                self.refresh_data()
                messagebox.showinfo("Thành công", f"Đã nhập {count} tài khoản từ Excel.")
                
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không đọc được file: {e}")

    def export_cookies_xlsx(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Thiếu thư viện", "Vui lòng cài đặt openpyxl: pip install openpyxl")
            return
            
        fp = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if fp:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                # include proxy column
                ws.append(["c_user", "xs", "proxy"])

                for acc in self.all_accounts:
                    ws.append([acc['c_user'], acc['xs'], acc.get('proxy', '')])
                
                wb.save(fp)
                messagebox.showinfo("Thành công", "Đã xuất file Excel (.xlsx).")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không ghi được file: {e}")

    def export_history_xlsx(self):
        if not self.history_rows:
            return
            
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Thiếu thư viện", "Vui lòng cài đặt openpyxl: pip install openpyxl")
            return

        fp = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if fp:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                # Header
                if hasattr(self, 'history_header'):
                    ws.append(self.history_header)
                else:
                    ws.append(["Thời gian", "URL", "Hạng mục", "Chi tiết", "Kết quả", "Account"])
                # Rows
                for row in self.history_rows:
                    ws.append(row)
                wb.save(fp)
                messagebox.showinfo("Thành công", "Đã xuất lịch sử ra Excel.")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không ghi được file: {e}")

    def add_report_set(self):
        cat = self.combo_category.get()
        detail = self.combo_detail.get()
        if cat and detail:
            self.report_sets.append((cat, detail))
            self.list_report_sets.insert('end', f"{cat} -> {detail}")
        else:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng chọn Hạng mục và Chi tiết trước khi thêm.")

    def clear_report_set(self):
        self.report_sets = []
        self.list_report_sets.delete(0, 'end')

    # --- BATCH LOGIC ---
    def start_batch(self):
        url = self.entry_url.get().strip()
        if not url:
            messagebox.showwarning("Thiếu URL", "Vui lòng nhập URL.")
            return
            
        # Use all_accounts instead of tree items
        if not self.all_accounts:
            messagebox.showwarning("Thiếu Account", "Danh sách tài khoản trống.")
            return

        self.is_running = True
        self.stop_event.clear()
        self.btn_run.config(state='disabled')
        self.btn_stop.config(state='normal')
        
        # Reset status for all accounts
        for acc in self.all_accounts:
            acc['status'] = "Chờ"
            acc['result'] = ""
        self.refresh_data()

        # Initialize batch timing and counters
        self.batch_start_time = time.time()
        with self.lock:
            self.processed_count = 0
            self.success_count = 0
            self.failure_count = 0
            self.total_accounts_in_batch = len(self.all_accounts)

        # Start periodic stats updater
        self.update_batch_stats()
            
        # Queue
        q = queue.Queue()
        for acc in self.all_accounts:
            # enqueue id, cookie, and per-account proxy (if any)
            q.put((acc['id'], acc['cookie'], acc.get('proxy', '').strip()))
            
        # Threads
        try:
            num_threads = int(self.spin_threads.get())
        except:
            num_threads = 1
            
        # Prepare report sets
        current_sets = list(self.report_sets)
        # If empty, use the currently selected one
        if not current_sets:
            cat = self.combo_category.get()
            detail = self.combo_detail.get()
            if cat and detail:
                current_sets.append((cat, detail))
            else:
                messagebox.showwarning("Thiếu cấu hình", "Vui lòng chọn Hạng mục báo cáo hoặc thêm vào bộ báo cáo.")
                self.btn_run.config(state='normal')
                self.btn_stop.config(state='disabled')
                return

        threading.Thread(target=self.run_queue, args=(q, num_threads, url, current_sets)).start()

    def run_queue(self, q, num_threads, url, report_sets):
        proxy = self.entry_proxy.get().strip()
        headless = self.var_headless.get()
        
        def worker():
            while not self.stop_event.is_set():
                try:
                    item_id, cookie, acc_proxy = q.get(timeout=1)
                except queue.Empty:
                    break
                
                # Pick random report config
                if report_sets:
                    r_cat, r_detail = random.choice(report_sets)
                    # Use per-account proxy if provided, otherwise use global proxy
                    use_proxy = acc_proxy if acc_proxy else proxy
                    self.process_one_account(item_id, cookie, url, r_cat, r_detail, use_proxy, headless)
                
                q.task_done()

        threads = []
        for _ in range(num_threads):
            t = threading.Thread(target=worker)
            t.start()
            threads.append(t)
            
        for t in threads:
            t.join()
            
        self.root.after(0, self.on_batch_finished)

    def process_one_account(self, item_id, cookie, url, cat, detail, proxy, headless):
        # Reduce UI updates to essential states to save overhead
        status_msg = "Đang chạy..."
        if proxy:
            # Show short proxy info
            try:
                # If user:pass@host:port -> host
                # If host:port -> host
                p_clean = proxy.strip()
                if "@" in p_clean:
                    host = p_clean.split("@")[1].split(":")[0]
                else:
                    host = p_clean.split(":")[0]
                status_msg = f"Run ({host})..."
            except:
                status_msg = "Run (Proxy)..."
        
        self.update_item(item_id, "status", status_msg)
        
        # Extract c_user for logging
        c_user = ""
        try:
            if "c_user=" in cookie:
                c_user = cookie.split("c_user=")[1].split(";")[0]
        except:
            pass

        success_flag = False
        bm = BrowserManager()
        with self.lock:
            self.active_browsers[item_id] = bm
            
        try:
            ok, msg = bm.start_browser(proxy, headless=headless)
            if not ok:
                self.update_item(item_id, "status", "Lỗi Start")
                self.update_item(item_id, "result", msg)
                log_report(url, cat, detail, f"Start Failed: {msg}", c_user)
                return

            # Skip "Inject Cookie..." update to reduce UI lag
            # self.update_item(item_id, "status", "Inject Cookie...")
            bm.inject_cookies(cookie)
            
            # Skip "Đang báo cáo..." update
            # self.update_item(item_id, "status", "Đang báo cáo...")
            ok_report, msg_report = bm.navigate_and_report(url, cat, detail)
            
            if ok_report:
                success_flag = True
                self.update_item(item_id, "status", "Hoàn thành")
                self.update_item(item_id, "result", msg_report)
                log_report(url, cat, detail, "Success", c_user)
            else:
                self.update_item(item_id, "status", "Lỗi")
                self.update_item(item_id, "result", msg_report)
                log_report(url, cat, detail, f"Failed: {msg_report}", c_user)
            
            # Final screenshot
            b64 = bm.get_screenshot_base64()
            self.save_screenshot_to_disk(item_id, b64)
            
        except Exception as e:
            self.update_item(item_id, "status", "Lỗi")
            self.update_item(item_id, "result", str(e))
            log_report(url, cat, detail, f"Exception: {str(e)}", c_user)
            
            # Error screenshot
            b64 = bm.get_screenshot_base64()
            self.save_screenshot_to_disk(item_id, b64)
        finally:
            bm.close()
            with self.lock:
                if item_id in self.active_browsers:
                    del self.active_browsers[item_id]
                # Update counters
                try:
                    self.processed_count += 1
                    if success_flag:
                        self.success_count += 1
                    else:
                        self.failure_count += 1
                except Exception:
                    pass

            # Request UI update of stats in main thread
            try:
                self.root.after(0, self.update_batch_stats)
            except Exception:
                pass

    def update_item(self, item_id, col, val):
        # Update Data Model (O(1))
        acc = self.account_map.get(item_id)
        if acc:
            acc[col] = val
            
        # Update UI only if visible
        # Check if item_id is in current treeview children
        if self.tree.exists(item_id):
            self.ui_queue.put((self.tree.set, (item_id, col, val)))

    def save_screenshot_to_disk(self, item_id, b64_data):
        if not b64_data: return
        try:
            path = os.path.join(self.temp_dir, f"{item_id}.png")
            with open(path, "wb") as f:
                f.write(base64.b64decode(b64_data))
        except Exception:
            pass

    def get_screenshot_from_disk(self, item_id):
        path = os.path.join(self.temp_dir, f"{item_id}.png")
        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    return base64.b64encode(f.read()).decode('utf-8')
            except:
                return None
        return None

    def stop_batch(self):
        self.stop_event.set()
        self.btn_stop.config(state='disabled')
        self.lbl_status.config(text="Đang dừng...")

    def on_batch_finished(self):
        self.is_running = False
        self.btn_run.config(state='normal')
        self.btn_stop.config(state='disabled')
        self.lbl_status.config(text="Đã hoàn tất batch.")
        # Final stats update
        try:
            self.update_batch_stats()
        except Exception:
            pass

        messagebox.showinfo("Xong", "Đã chạy xong danh sách.")

    def format_seconds(self, sec):
        try:
            sec = int(sec)
            h = sec // 3600
            m = (sec % 3600) // 60
            s = sec % 60
            return f"{h:02d}:{m:02d}:{s:02d}"
        except Exception:
            return "--:--:--"

    def update_batch_stats(self):
        # Update elapsed, remaining, success % labels. Must be called from main thread.
        if not self.batch_start_time:
            # Nothing to show
            self.lbl_elapsed.config(text="Elapsed: 00:00:00")
            self.lbl_remaining.config(text="Remaining: --:--:--")
            self.lbl_success.config(text=f"Success: 0% (0/0)")
            return

        elapsed = time.time() - self.batch_start_time

        with self.lock:
            processed = getattr(self, 'processed_count', 0)
            success = getattr(self, 'success_count', 0)
            total = getattr(self, 'total_accounts_in_batch', 0)

        # Completion % (processed / total)
        completion_pct = 0.0
        if total > 0:
            completion_pct = (processed / total) * 100.0

        # Success % among processed
        success_pct = 0.0
        if processed > 0:
            success_pct = (success / processed) * 100.0

        # Estimate remaining time using average
        remaining_text = "--:--:--"
        try:
            if processed > 0 and processed < total:
                avg = elapsed / processed
                remain = avg * (total - processed)
                remaining_text = self.format_seconds(remain)
            elif processed >= total:
                remaining_text = "00:00:00"
        except Exception:
            remaining_text = "--:--:--"

        # Update labels
        try:
            self.lbl_elapsed.config(text=f"Elapsed: {self.format_seconds(elapsed)}")
            self.lbl_remaining.config(text=f"Remaining: {remaining_text}")
            self.lbl_success.config(text=f"Success: {int(success_pct)}% ({success}/{processed if processed>0 else 0}) | {int(completion_pct)}% done")
        except Exception:
            pass

        # Schedule next update while running
        if self.is_running:
            try:
                self.root.after(1000, self.update_batch_stats)
            except Exception:
                pass